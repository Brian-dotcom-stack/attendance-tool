"""
excel_updater.py

Writes absence status codes into the monthly Excel attendance sheets.

Key design decisions:
- Names are read from the 'Employee Master' sheet (plain text values).
  The monthly sheets use =IF() formulas in col B which openpyxl reads as
  formula strings, not evaluated values, when loaded with data_only=False.
  The Employee Master sheet has literal strings and the row numbers are
  identical across all sheets, so we build the map once from there.
- Employee map is built once per call, not per day.
- Matching: exact -> substring -> fuzzy (token_set_ratio >= 80).
"""

from datetime import datetime, date, timedelta
from typing import Dict, List, Optional
import openpyxl
import re
from rapidfuzz import process, fuzz

MONTH_SHEETS = {
    1: "Jan",  2: "Feb",  3: "Mar",  4: "Apr",
    5: "May",  6: "Jun",  7: "Jul",  8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

EMPLOYEE_MASTER_SHEET = "Employee Master"
EMPLOYEE_START_ROW    = 8
NAME_COL              = 2   # col B
DAY_1_COL             = 4   # col D = day 1


def normalize(name: str) -> str:
    name = str(name)
    name = name.replace("\t", " ").replace("\n", " ")
    name = re.sub(r"\s+", " ", name)
    return name.strip().lower()


def to_date(value) -> Optional[date]:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:10]).date()
        except (ValueError, TypeError):
            return None
    return None


def day_column(day: int) -> int:
    return DAY_1_COL + day - 1


def build_employee_map(wb: openpyxl.Workbook) -> Dict[str, int]:
    """
    Read employee names from 'Employee Master' (literal text, not formulas).
    Row numbers are identical in every monthly sheet.
    """
    if EMPLOYEE_MASTER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{EMPLOYEE_MASTER_SHEET}' not found. Available: {wb.sheetnames}")

    sheet = wb[EMPLOYEE_MASTER_SHEET]
    emp_map: Dict[str, int] = {}

    for row in range(EMPLOYEE_START_ROW, sheet.max_row + 1):
        value = sheet.cell(row=row, column=NAME_COL).value
        if not value:
            continue
        name = normalize(value)
        if name in ("employee name", "reported to", ""):
            continue
        emp_map[name] = row

    return emp_map


def match_name(name: str, candidates: Dict[str, int]) -> Optional[int]:
    """
    1. Exact match
    2. Substring  ("endurance" in "endurance iziegbe osarumwense")
    3. Fuzzy      (token_set_ratio >= 80)
    """
    if not candidates:
        return None

    name_clean = normalize(name)

    if name_clean in candidates:
        return candidates[name_clean]

    for key, row in candidates.items():
        if name_clean in key or key in name_clean:
            return row

    match = process.extractOne(name_clean, list(candidates.keys()), scorer=fuzz.token_set_ratio)
    if match and match[1] >= 80:
        return candidates[match[0]]

    return None


def update_excel(
    excel_path: str,
    output_path: str,
    cancellations: List[Dict],
    default_status: str = "UA",
) -> Dict:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    emp_map = build_employee_map(wb)

    stats: Dict = {"updated": 0, "unmatched": 0, "unmatched_names": []}

    for record in cancellations:
        name   = record.get("name")
        start  = to_date(record.get("start_date") or record.get("start"))
        end    = to_date(record.get("end_date")   or record.get("end"))
        status = record.get("excel_status") or default_status

        if not name or not start or not end:
            continue

        row = match_name(name, emp_map)

        if not row:
            stats["unmatched"] += 1
            if name not in stats["unmatched_names"]:
                stats["unmatched_names"].append(name)
            continue

        current = start
        while current <= end:
            sheet_name = MONTH_SHEETS.get(current.month)
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                col = day_column(current.day)
                sheet.cell(row=row, column=col).value = status
                stats["updated"] += 1
            current += timedelta(days=1)

    wb.save(output_path)
    return stats